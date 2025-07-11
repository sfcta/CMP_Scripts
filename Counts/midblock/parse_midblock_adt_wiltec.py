import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import polars as pl
from openpyxl import load_workbook


def read_day_dir_dayofweek(workbook, sheet_name, day_num):
    row = 8 + day_num * 48
    date = workbook[sheet_name][f"D{row}"].value
    if date is None:
        raise EOFError(
            "No data here; day_num is probably beyond "
            "the number of days collected at this location."
        )
    else:
        return date.split(" ")[0]


def is_weekday(dayofweek: str):
    if dayofweek in {"MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"}:
        return True
    elif dayofweek in {"SATURDAY", "SUNDAY"}:
        return False
    else:
        raise ValueError(f"{dayofweek} unrecognized.")


def is_tuetothu(dayofweek: str):
    if dayofweek in {"TUESDAY", "WEDNESDAY", "THURSDAY"}:
        return True
    elif dayofweek in {"MONDAY", "FRIDAY", "SATURDAY", "SUNDAY"}:
        return False
    else:
        raise ValueError(f"{dayofweek} unrecognized.")


def read_day_dir_counts_df(filepath, sheet_name, direction, day_num):
    """Read the counts table from a single day and direction

    Parameters
    ----------
    filepath : _type_
        _description_
    sheet_name : _type_
        _description_
    direction : _type_
        _description_
    day_num : _type_
        the nth day of data collection within this worksheet

    Returns
    -------
    _type_
        _description_

    Raises
    ------
    ValueError
        _description_
    """
    if direction == 0:
        cols = "A:F"
    elif direction == 1:
        cols = "H:M"
    else:
        raise ValueError("direction must be 0 or 1")
    header = 10 + day_num * 48
    try:
        df = pd.read_excel(
            filepath,
            sheet_name,
            usecols=cols,
            header=header,
            skiprows=[header + 1],
            nrows=24,
        )
    except ValueError:
        raise EOFError(
            "Reached end of sheet; day_num is probably beyond "
            "the number of days collected at this location."
        )
    if direction == 1:
        # pandas append ".1" to column names to the right table (direction 1)
        # since they duplicate the headers of the left table (direction 0)
        df.columns = df.columns.str.rstrip(".1")
    return df.rename(columns={"    TIME": "TIME", "HOUR ": "HOUR_TOTALS"}).set_index(
        "TIME"
    )


def get_am_peak_totals(day_dir_counts_df):
    """Get a.m. peak (0700-0900) totals for a single day and direction

    Parameters
    ----------
    day_dir_counts_df : _type_
        the counts table from a single day and direction

    Returns
    -------
    _type_
        _description_
    """
    return day_dir_counts_df.iloc[7:9]["HOUR_TOTALS"].sum()


def get_pm_peak_totals(day_dir_counts_df):
    """Get p.m. peak (1630-1830) totals for a single day and direction

    Parameters
    ----------
    day_dir_counts_df : _type_
        the counts table from a single day and direction

    Returns
    -------
    _type_
        _description_
    """
    return np.sum(
        (
            day_dir_counts_df.iloc[16][["30-45", "45-60"]].sum(),
            day_dir_counts_df.iloc[17]["HOUR_TOTALS"],
            day_dir_counts_df.iloc[18][["00-15", "15-30"]].sum(),
        )
    )


def read_day_total(workbook, sheet_name, direction, day_num):
    if direction == 0:
        col = "F"
    elif direction == 1:
        col = "M"
    else:
        raise ValueError("direction must be 0 or 1")
    row = 37 + day_num * 48
    total = workbook[sheet_name][f"{col}{row}"].value
    if total is None:
        raise EOFError(
            "No data here; day_num is probably beyond "
            "the number of days collected at this location."
        )
    else:
        return total


def read_totals(filepath, workbook, sheet_name, direction, dow_mode):
    """_summary_

    Note that filepath and workbook are both needed because pandas is used to
    read the am/pm peak totals, while openpyxl is used to read the daily totals,
    and loading the openpyxl workbook every function run would be slow.

    Parameters
    ----------
    filepath : _type_
        _description_
    workbook : _type_
        _description_
    sheet_name : _type_
        _description_
    direction : _type_
        _description_

    Returns
    -------
    _type_
        _description_
    """
    if dow_mode not in {"tuetothu", "anyweekday", "anyday"}:
        raise KeyError("dow_mode should be tuetothu, anyweekday, or anyday")
    day_num = -1
    am_peak_totals = []
    pm_peak_totals = []
    daily_totals = []
    while True:
        day_num += 1
        try:
            dow = read_day_dir_dayofweek(workbook, sheet_name, day_num)
            if (
                (dow_mode == "anyday")
                or ((dow_mode == "anyweekday") and is_weekday(dow))
                or ((dow_mode == "tuetothu") and is_tuetothu(dow))
            ):
                df = read_day_dir_counts_df(filepath, sheet_name, direction, day_num)
                am_peak_totals.append(get_am_peak_totals(df))
                pm_peak_totals.append(get_pm_peak_totals(df))
                daily_totals.append(
                    read_day_total(workbook, sheet_name, direction, day_num)
                )
        except EOFError:  # assume we've reached the end of the sheet
            break
    return am_peak_totals, pm_peak_totals, daily_totals


def read_sheet_directions(workbook, sheet_name):
    return (
        workbook[sheet_name]["D10"].value,
        workbook[sheet_name]["K10"].value,
    )


def read_sheet_mean_totals(filepath, workbook, sheet_name, dow_mode, raw_values=False):
    """Read count totals (i.e. volumes) from a single worksheet

    This is where we start referring to totals as volumes

    Parameters
    ----------
    filepath : _type_
        _description_
    workbook : _type_
        _description_
    sheet_name : _type_
        _description_

    Returns
    -------
    _type_
        _description_
    """
    # direction of tables on the right
    dir0, dir1 = read_sheet_directions(workbook, sheet_name)
    # avg daily total (in each direction)
    vols_dir0 = map(
        np.mean,
        read_totals(filepath, workbook, sheet_name, 0, dow_mode),
    )
    rows = [[sheet_name, dir0, *vols_dir0]]
    if dir1 != "O":
        vols_dir1 = map(
            np.mean, read_totals(filepath, workbook, sheet_name, 1, dow_mode)
        )
        rows.append([sheet_name, dir1, *vols_dir1])
    if raw_values:
        return rows
    else:
        column_names = (
            "sheet_name",
            "direction",
            "cmp_am_peak_vol",
            "cmp_pm_peak_vol",
            "daily_vol",
        )
        return pd.DataFrame(
            rows,
            columns=column_names,
        )


def read_xlsxs_volumes(xlsx_filepaths, dow_mode):
    return pl.concat(read_xlsx_volumes(f, dow_mode) for f in xlsx_filepaths)


def read_xlsx_volumes(filepath, dow_mode):
    print("parsing file:", filepath)
    wb = load_workbook(filename=filepath, read_only=True)
    column_names = (
        "sheet_name",
        "direction",
        "cmp_am_peak_vol",
        "cmp_pm_peak_vol",
        "daily_vol",
    )
    count_totals = []
    sheet_names = [sn for sn in wb.sheetnames if not sn.lower().endswith("photo")]
    for sheet_name in sheet_names:
        print("parsing sheet:", sheet_name)
        count_totals.extend(
            read_sheet_mean_totals(
                filepath,
                wb,
                sheet_name,
                dow_mode,
                raw_values=True,
            )
        )
    return pl.DataFrame(count_totals, schema=column_names, orient="row")


def add_cmp_non_peak_volume(volumes_df):
    return volumes_df.with_columns(
        cmp_non_peak_vol=(
            pl.col("daily_vol") - pl.col("cmp_am_peak_vol") - pl.col("cmp_pm_peak_vol")
        )
    )


def rename_sheet_name_to_location(volumes_df, locations_filepath):
    locations = pl.read_csv(
        locations_filepath,
        columns=["id_nodir_2023", "id_withdir_2023", "name_2023", "direction"],
    ).rename(
        {
            "id_nodir_2023": "location_id_nodir_2023",
            "id_withdir_2023": "location_id_withdir_2023",
            "name_2023": "location",
        }
    )
    return (
        volumes_df.with_columns(
            location_id_nodir_2023=pl.col("sheet_name")
            .str.replace(" ", "-")
            # str.split can't handle regex as of polars v1
            # r"[-\\s]+"  # split on "-" or " "
            .str.split("-")
            .list.get(1)
            .cast(int)
        )
        .join(locations, on=["location_id_nodir_2023", "direction"])
        .select(  # reorder columns and drop sheet_name
            "location_id_nodir_2023",
            "location_id_withdir_2023",
            "location",
            "direction",
            "cmp_am_peak_vol",
            "cmp_pm_peak_vol",
            "cmp_non_peak_vol",
            "daily_vol",
        )
    )


def parse_wiltec_excel(
    raw_xlsx_filepaths, locations_filepath, output_csv_filepath_stem, dow_mode
):
    output_csv_filepath = f"{output_csv_filepath_stem}-{dow_mode}.csv"
    print("day of week mode:", dow_mode)
    print("parsing workbooks", list(raw_xlsx_filepaths))
    print("output filepath:", output_csv_filepath)
    rename_sheet_name_to_location(
        add_cmp_non_peak_volume(read_xlsxs_volumes(raw_xlsx_filepaths, dow_mode)),
        locations_filepath,
    ).write_csv(output_csv_filepath)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "raw_xlsx_glob_filepaths", help="Wiltec-collected xlsx input (glob) filpath(s)"
    )
    parser.add_argument(
        "locations_filepath",
        help=(
            "Locations xlsx input filepath: "
            "check location IDs and raw_xlsx sheet numbers match"
        ),
    )
    parser.add_argument("output_csv_filepath_stem", help="Output CSV filepath stem")
    args = parser.parse_args()

    dow_mode = "tuetothu"  # day of week mode: tuetothu, anyweekday, anyday
    # anyweekday was used for CMP 2023, tuetothu used from CMP2025 onwards

    raw_xlsx_glob_filepaths = Path(args.raw_xlsx_glob_filepaths)
    raw_xlsx_filepaths = list(
        raw_xlsx_glob_filepaths.parent.glob(raw_xlsx_glob_filepaths.name)
    )
    parse_wiltec_excel(
        raw_xlsx_filepaths,
        args.locations_filepath,
        args.output_csv_filepath_stem,
        dow_mode,
    )
